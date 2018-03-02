import datetime
import hashlib
import re
from pathlib import Path
from zipfile import ZipFile

from django.utils.translation import ugettext as _
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from apps.core.models import OmicsArea, OmicsUnitType, Strain
from apps.data.models import Repository
from .. import exceptions


def style_range(ws,
                cell_range,
                border=Border(),
                fill=None,
                font=None,
                alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    This snippet has been inspired by
    `openpyxl documentation <http://openpyxl.readthedocs.io/en/default/styles.html#styling-merged-cells>`_.

    Parameters
    ----------
    ws :  :obj:`Worksheet`
        Excel worksheet instance
    cell_range : str
        An excel range to style (e.g. A1:F20)
    border : :obj:`Border`, optional
        An openpyxl Border
    fill : :obj:`PatternFill` or :obj:`GradientFill`, optional
        An openpyxl PatternFill or GradientFill
    font : :obj:`Font`, optional
        An openpyxl Font object
    alignment : :obj:`Alignment`, optional
        An openpyxl Alignment object
    """  # noqa

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    rows = ws[cell_range]
    for cell in rows[0]:
        cell.border += top
    for cell in rows[-1]:
        cell.border += bottom

    for row in rows:
        left_cell = row[0]
        right_cell = row[-1]
        left_cell.border += left
        right_cell.border += right
        if fill:
            for cell in row:
                cell.fill = fill
        if font:
            for cell in row:
                cell.font = font
        if alignment:
            for cell in row:
                cell.alignment = alignment


def sha256_checksum(filename):
    """
    Calculate a sha256 checksum for a file

    Parameters
    ----------
    filename : str
        Path to the sha256 checksum target file

    Returns
    -------
    checksum : str
        Target file sha256 checksum (hexdigest)
    """
    m = hashlib.sha256()
    with open(filename, 'rb') as f:
        while True:
            chunk = f.read(4096)
            if not chunk:
                break
            m.update(chunk)
    return m.hexdigest()


def get_template_version(xlsx_filename):
    """
    Calculate template version based on CRC-32 of xlsx zip archive files

    Parameters
    ----------
    filename : str
        Path to the xlsx target file

    Returns
    -------
    version : str
        Target file version (hexdigest)
    """
    z = ZipFile(xlsx_filename)
    m = hashlib.sha256()
    # Concatenate all files CRCs to calculate a hash
    m.update(b''.join(bytes(str(zi.CRC), 'utf-8') for zi in z.infolist()))
    return m.hexdigest()


def generate_template(filename):
    """
    Generate Pixel XLS template that will be used during the imporation
    process.

    Parameters
    ----------
    filename : str
        Pixel's XLSX template file name

    Returns
    -------
    checksum : str
        Pixel's XLSX template file checksum
    version : str
        Pixel's XLSX template file version
    """

    wb = Workbook()
    ws = wb.active
    ws.title = _("Import information for Pixel")

    # Styles
    # https://coolors.co/4c5760-c5e8cd-fdffd9-ffd7ad-d1a690
    section_title_font = Font(bold=True, size=18, color='d1a690')
    field_font = Font(color='4c5760')
    required_field_font = Font(bold=True, color='4c5760')
    comment_fill = PatternFill('solid', fgColor='c5e8cd')
    pixelset_header_fill = PatternFill('solid', fgColor='ffd7ad')
    long_text_align = Alignment(vertical='center', wrap_text=True)
    user_data_fill = PatternFill('solid', fgColor='fdffd9')

    # Validators
    def get_node_repr(node):
        return '{}{}{}'.format(
            '—' * node.level,
            ' ' if node.level else '',
            node.name
        )

    omics_areas_names = ','.join(
        get_node_repr(oa) for oa in OmicsArea.objects.all()
    )
    omics_areas_validator = DataValidation(
        type='list',
        formula1='"{}"'.format(omics_areas_names),
        allow_blank=False
    )
    omics_areas_validator.error = _("Omics Area does not exists")
    omics_areas_validator.prompt = _("Select an Omics Area")
    ws.add_data_validation(omics_areas_validator)

    data_source_validator = DataValidation(
        type='list',
        formula1='"Published, Unpublished"',
        allow_blank=False
    )
    data_source_validator.error = _("Value not allowed")
    data_source_validator.prompt = _("Select a data source")
    ws.add_data_validation(data_source_validator)

    omics_unit_type_names = ','.join(
        t.name for t in OmicsUnitType.objects.all()
    )
    omics_unit_type_validator = DataValidation(
        type='list',
        formula1='"{}"'.format(omics_unit_type_names),
        allow_blank=False
    )
    omics_unit_type_validator.error = _("Value not allowed")
    omics_unit_type_validator.prompt = _("Select a type of omics unit")
    ws.add_data_validation(omics_unit_type_validator)

    strain_names = ','.join(
        '{} ({})'.format(s.name, s.species.name) for s in Strain.objects.all()
    )
    strain_validator = DataValidation(
        type='list',
        formula1='"{}"'.format(strain_names),
        allow_blank=False
    )
    strain_validator.error = _("Value not allowed")
    strain_validator.prompt = _("Select a strain")
    ws.add_data_validation(strain_validator)

    # Ranges
    comment_range = 'A{0}:K{0}'

    # Experiment
    ws['A1'] = _("Experiment")
    ws['A1'].font = section_title_font
    ws.row_dimensions[1].height = 40

    ws.merge_cells(comment_range.format(2))
    ws['A2'] = _(
        "# This section describes the experimental conditions that were "
        "applied to obtain the secondary datafile (see section 'Analysis' "
        "below). Note that these experiments can be already published (in "
        "this situation a DOI is required) or not (in this situation a "
        "laboratory has to be specified)."
    )
    ws['A2'].fill = comment_fill
    ws.row_dimensions[2].height = 40
    style_range(
        ws,
        comment_range.format(2),
        fill=comment_fill,
        alignment=long_text_align
    )

    ws['A3'] = _("Omics area")
    ws['A3'].font = required_field_font
    ws['B3'].fill = user_data_fill
    omics_areas_validator.add(ws['B3'])

    ws['A4'] = _("Completion date")
    ws['A4'].font = field_font
    ws['B4'].fill = user_data_fill

    ws['A5'] = _("Summary")
    ws['A5'].font = field_font
    ws['B5'].fill = user_data_fill

    ws['A6'] = _("Release date")
    ws['A6'].font = field_font
    ws['B6'].fill = user_data_fill

    ws['A7'] = _("Data source")
    ws['A7'].font = field_font
    ws['B7'].fill = user_data_fill
    data_source_validator.add(ws['B7'])

    ws['A8'] = _("Reference (entry)")
    ws['A8'].font = field_font
    ws['B8'].fill = user_data_fill
    ws['B8'].comment = Comment(
        _("If this work has been published, we expect a DOI in this cell."),
        _("Pixel's administrator")
    )

    # Analysis
    ws['A10'] = _("Analysis")
    ws['A10'].font = section_title_font
    ws.row_dimensions[10].height = 40

    ws.merge_cells(comment_range.format(11))
    ws['A11'] = _(
        "# This section describes the data analyses that were performed on "
        "secondary datasets to obtain pixel datasets. The secondary "
        "datafile has to be associated to the pixel datasets during the "
        "import process."
    )
    ws['A11'].fill = comment_fill
    ws.row_dimensions[11].height = 40
    style_range(
        ws,
        comment_range.format(11),
        fill=comment_fill,
        alignment=long_text_align
    )

    ws['A12'] = _("Name of secondary data file")
    ws['A12'].font = required_field_font
    ws['B12'].fill = user_data_fill

    ws['A13'] = _("Name of notebook file")
    ws['A13'].font = field_font
    ws['B13'].fill = user_data_fill

    ws['A14'] = _("Description")
    ws['A14'].font = field_font
    ws['B14'].fill = user_data_fill

    ws['A15'] = _("Date of the analysis")
    ws['A15'].font = field_font
    ws['B15'].fill = user_data_fill

    # Pixel datasets
    ws['A17'] = _("Pixel datasets")
    ws['A17'].font = section_title_font

    ws.merge_cells(comment_range.format(18))
    ws['A18'] = _(
        "# This section lists and describes each pixel datasets to be "
        "imported in the system. These files have to be associated to the "
        "secondary datafile (and the notebook datafile if available) "
        "during the import process. A specific comment can be added for "
        "each set of Pixel to better describe their differences."
    )
    ws['A18'].fill = comment_fill
    ws.row_dimensions[18].height = 40
    style_range(
        ws,
        comment_range.format(18),
        fill=comment_fill,
        alignment=long_text_align
    )

    style_range(ws, comment_range.format(19), fill=pixelset_header_fill)
    ws['A19'] = _("File name")
    ws['B19'] = _("Omics Unit type")
    ws['C19'] = _("Strain (Species)")
    ws['D19'] = _("Comment")

    for row in range(20, 31, 1):
        omics_unit_type_validator.add(ws['B{}'.format(row)])
        for column in ('A', 'B', 'C', 'D'):
            ws[f'{column}{row}'].fill = user_data_fill
        strain_validator.add(ws['C{}'.format(row)])

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    wb.save(filename)

    return (
        sha256_checksum(filename),
        get_template_version(filename)
    )


def parse_template(meta_path, serialized=False):

    wb = load_workbook(meta_path)
    ws = wb.active

    meta = {
        'experiment': {
            'omics_area': None,
            'completion_date': None,
            'summary': None,
            'release_date': None,
            'repository': None,
            'entry': None,
        },
        'analysis': {
            'secondary_data_path': None,
            'notebook_path': None,
            'description': None,
            'date': None,
        },
        'datasets': [
            # [path, omics_unit_type, strain, comment]
        ]
    }

    # -- Experiment
    # Omics area
    omics_area = ws['B3'].value.lstrip('—').strip()
    if not len(omics_area):
        raise exceptions.MetaFileParsingError(
            _("An omics area is required (B3)")
        )
    qs = OmicsArea.objects.filter(name=omics_area)
    if qs.count() != 1:
        raise exceptions.MetaFileParsingError(
            _("Selected omics area {} does not exists yet").format(
                omics_area
            )
        )
    omics_area = qs.get()
    if serialized:
        omics_area = omics_area.name
    meta['experiment']['omics_area'] = omics_area

    # Misc
    try:
        completion_date = datetime.date(year=ws['B4'].value, month=1, day=1)
    except TypeError:
        raise exceptions.MetaFileParsingError(
            _("Completion date should be a single integer, e.g. 2017")
        )
    meta['experiment']['completion_date'] = completion_date
    meta['experiment']['summary'] = ws['B5'].value.strip()
    try:
        release_date = datetime.date(year=ws['B6'].value, month=1, day=1)
    except TypeError:
        raise exceptions.MetaFileParsingError(
            _("Experiment release date should be a single integer, e.g. 2017")
        )
    meta['experiment']['release_date'] = release_date

    # Repository & Entry
    data_source = ws['B7'].value.strip()
    reference = ws['B8'].value.strip()
    if not len(data_source):
        raise exceptions.MetaFileParsingError(
            _("A data source is required (B7)")
        )
    if not len(reference):
        raise exceptions.MetaFileParsingError(
            _("A reference is required (B8)")
        )
    if data_source not in ('Published', 'Unpublished'):
        raise exceptions.MetaFileParsingError(
            _("{} is not a valid data source").format(data_source)
        )
    if data_source == 'Published':
        qs = Repository.objects.filter(name='PUBMED')
    elif data_source == 'Unpublished':
        qs = Repository.objects.filter(name='PARTNERS')
    if qs.count() != 1:
        raise exceptions.MetaFileParsingError(
            _("Data source has no corresponding repository")
        )
    repository = qs.get()
    if serialized:
        repository = repository.name
    meta['experiment']['repository'] = repository
    meta['experiment']['entry'] = reference

    # -- Analysis
    # Secondary data
    secondary_data_path = meta_path.parent / Path(ws['B12'].value)
    if not secondary_data_path.exists():
        raise exceptions.MetaFileParsingError(
            _("Secondary data file {} not found").format(
                secondary_data_path.name
            )
        )
    if serialized:
        secondary_data_path = secondary_data_path.name
    meta['analysis']['secondary_data_path'] = secondary_data_path

    # Notebook
    notebook_path = meta_path.parent / Path(ws['B13'].value)
    if not notebook_path.exists():
        raise exceptions.MetaFileParsingError(
            _("Notebook file {} not found").format(notebook_path.name)
        )
    if serialized:
        notebook_path = notebook_path.name
    meta['analysis']['notebook_path'] = notebook_path

    # Misc
    meta['analysis']['description'] = ws['B14'].value.strip()

    try:
        analysis_date = datetime.date(year=ws['B15'].value, month=1, day=1)
    except TypeError:
        raise exceptions.MetaFileParsingError(
            _("Analysis date should be a single integer, e.g. 2017")
        )
    meta['analysis']['date'] = analysis_date

    # -- Pixels
    for row in range(20, 31, 1):

        # Skip empty rows
        cells = ['{}{}'.format(c, row) for c in ('A', 'B', 'C', 'D')]
        if not all(ws[cell].value for cell in cells):
            continue

        dataset_path = meta_path.parent / Path(ws['A{}'.format(row)].value)
        if not dataset_path.exists():
            raise exceptions.MetaFileParsingError(
                _("Dataset file {} not found").format(dataset_path.name)
            )
        if serialized:
            dataset_path = dataset_path.name

        omics_unit_type_name = ws['B{}'.format(row)].value
        qs = OmicsUnitType.objects.filter(name=omics_unit_type_name)
        if qs.count() != 1:
            raise exceptions.MetaFileParsingError(
                _("Unknown OmicsUnit type {}").format(omics_unit_type_name)
            )
        omics_unit_type = qs.get()
        if serialized:
            omics_unit_type = omics_unit_type.name

        strain_name = ws['C{}'.format(row)].value
        m = re.match('(.+) \((.+)\)', strain_name)
        if not m:
            raise exceptions.MetaFileParsingError(
                _("Invalid strain/species pattern: {}").format(strain_name)
            )
        strain_name, species_name = m.groups()
        qs = Strain.objects.filter(
            name=strain_name,
            species__name=species_name
        )
        if qs.count() != 1:
            raise exceptions.MetaFileParsingError(
                _("Unknown strain: {} (species: {})").format(
                    strain_name, species_name
                )
            )
        strain = qs.get()
        if serialized:
            strain = strain.name

        comment = ws['D{}'.format(row)].value.strip()
        if not len(comment):
            raise exceptions.MetaFileParsingError(
                _("You need to add a comment for the {} dataset").format(
                    dataset_path.name
                )
            )

        meta['datasets'].append(
            [dataset_path, omics_unit_type, strain, comment]
        )

    return meta
