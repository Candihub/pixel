import datetime
from pathlib import Path

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment, Border, Color, Font, PatternFill, Side, colors
)

from apps.core.models import OmicsArea, OmicsUnitType, Strain
from apps.core.tests import CoreFixturesTestCase
from apps.data.models import Repository
from apps.submission.io.xlsx import (
    generate_template, get_template_version, parse_template, sha256_checksum,
    style_range
)


def test_style_range():

    wb = Workbook()
    ws = wb.active
    start = 1
    stop = 10
    column = 'A'
    cell_range = f'{column}{start}:{column}{stop}'

    # Border
    dashed_bottom_border = Border(bottom=Side(style='dashed'))
    style_range(ws, cell_range, border=dashed_bottom_border)
    cell = f'{column}{start}'
    assert ws[cell].border.top == dashed_bottom_border.top
    cell = f'{column}{stop}'
    assert ws[cell].border.bottom == dashed_bottom_border.bottom

    # Fill
    red_fill = PatternFill('solid', fgColor=colors.RED)
    style_range(ws, cell_range, fill=red_fill)
    for row in range(start, stop + 1):
        cell = f'{column}{row}'
        assert ws[cell].fill == red_fill

    # Font
    blue_font = Font(color=colors.BLUE)
    style_range(ws, cell_range, font=blue_font)
    for row in range(start, stop + 1):
        cell = f'{column}{row}'
        assert ws[cell].font == blue_font

    # Alignment
    center_vertical_alignment = Alignment(vertical='center')
    style_range(ws, cell_range, alignment=center_vertical_alignment)
    for row in range(start, stop + 1):
        cell = f'{column}{row}'
        assert ws[cell].alignment == center_vertical_alignment


def test_sha256_checksum():

    meta_filename = Path('apps/submission/fixtures/meta.xlsx')
    checksum = sha256_checksum(meta_filename)
    expected = (
        'd7db1382d72db9bb5611b385d5fc8fde7321ff0c1229a012f58e79c610104a48'
    )
    assert checksum == expected


def test_get_template_version():

    meta_filename = Path('apps/submission/fixtures/meta.xlsx')
    version = get_template_version(meta_filename)
    expected = (
        '6e13234a02f9225b4ca85e676ff20a4f7f1da0b3a7bf95fadf4c82989f78a249'
    )
    assert version == expected


class XLSXTemplateTestCase(CoreFixturesTestCase):

    @pytest.fixture(autouse=True)
    def set_tmpdir(self, tmpdir):
        self.tmpdir = tmpdir

    def test_generate_template(self):

        pixel_template = self.tmpdir.join('pixel-template.xlsx')
        checksum, version = generate_template(filename=pixel_template)

        wb = load_workbook(pixel_template)
        ws = wb.active

        # fields
        assert ws.title == 'Import information for Pixel'
        assert ws['A1'].value == 'Experiment'
        assert ws['A3'].value == 'Omics area'
        assert ws['A4'].value == 'Completion date'
        assert ws['A5'].value == 'Summary'
        assert ws['A6'].value == 'Release date'
        assert ws['A7'].value == 'Data source'
        assert ws['A8'].value == 'Reference (entry)'
        assert ws['A10'].value == 'Analysis'
        assert ws['A12'].value == 'Name of secondary data file'
        assert ws['A13'].value == 'Name of notebook file'
        assert ws['A14'].value == 'Description'
        assert ws['A15'].value == 'Date of the analysis'
        assert ws['A17'].value == 'Pixel datasets'
        assert ws['A19'].value == 'File name'
        assert ws['B19'].value == 'Omics Unit type'
        assert ws['C19'].value == 'Strain (Species)'
        assert ws['D19'].value == 'Comment'

        # user data
        user_data_cells = tuple(
            f'B{r}' for r in list(range(3, 9)) + list(range(12, 16))
        ) + tuple(
            f'{c}{r}' for r in range(20, 31) for c in ('A', 'B', 'C', 'D')
        )
        expected_user_data_fg_color = Color('fdffd9')
        for cell in user_data_cells:
            assert ws[cell].fill.fgColor == expected_user_data_fg_color


class ParseXLSXTemplateTestCase(CoreFixturesTestCase):

    def test_parse_template(self):

        meta_path = Path('apps/submission/fixtures/dataset-0001/meta.xlsx')
        meta = parse_template(meta_path)

        expected = OmicsArea.objects.get(name='Label free')
        assert meta['experiment']['omics_area'] == expected

        expected = datetime.date(year=2015, month=1, day=1)
        assert meta['experiment']['completion_date'] == expected

        expected = (
            'In these experiments, mass spectrometry analyses were performed '
            'in yeast Candida glabrata. Proteins were extracted using FASP '
            'protocol (by Camille Garcia from the platform proteomics@IJM). '
            'Technical and biolocal replicates were done in order to evaluate '
            'the variability associated to each type of data reproduction. '
            'Protein abundances were obtained with PROGENESIS software, '
            'following the standard procedure of the proteomics platform (in '
            '2015). Note that cell were submitted to an alkaline stress (1mL '
            'TRIS base), to observe modifications in protein abundances.'
        )
        assert meta['experiment']['summary'] == expected

        expected = datetime.date(year=2017, month=1, day=1)
        assert meta['experiment']['release_date'] == expected

        expected = Repository.objects.get(name='PARTNERS')
        assert meta['experiment']['repository'] == expected

        expected = 'Camadro laboratory'
        assert meta['experiment']['entry'] == expected

        expected = Path(
            'apps/submission/fixtures/dataset-0001/'
            '1503002-protein-measurements-PD2.1.csv'
        )
        assert meta['analysis']['secondary_data_path'] == expected

        expected = Path('apps/submission/fixtures/dataset-0001/NoteBook.R')
        assert meta['analysis']['notebook_path'] == expected

        expected = (
            'Protein abundances obtained in two cell growth conditions '
            '(alkaline pH or standard) were compared, in order to identify '
            'differentially expressed proteins. LIMMA method was applied with '
            'default parameters, in order to calculate p-values.'
        )
        assert meta['analysis']['description'] == expected

        expected = datetime.date(year=2017, month=1, day=1)
        assert meta['analysis']['date'] == expected

        expected = Path('apps/submission/fixtures/dataset-0001/Pixel_C10.txt')
        assert meta['datasets'][0][0] == expected

        protein = OmicsUnitType.objects.get(name='protein')
        assert meta['datasets'][0][1] == protein

        deltaHTU = Strain.objects.get(name='deltaHTU')
        assert meta['datasets'][0][2] == deltaHTU

        expected = (
            'This set of Pixel correspond to a time point T10 (10 minutes) '
            'after TRIS base addition in the cell growth media.'
        )
        assert meta['datasets'][0][3] == expected

        expected = Path('apps/submission/fixtures/dataset-0001/Pixel_C60.txt')
        assert meta['datasets'][1][0] == expected

        assert meta['datasets'][1][1] == protein

        assert meta['datasets'][1][2] == deltaHTU

        expected = (
            'This set of Pixel correspond to a time point T10 (10 minutes) '
            'after TRIS base addition in the cell growth media.'
        )
        assert meta['datasets'][1][3] == expected
